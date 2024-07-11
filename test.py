from openpyxl import load_workbook
from sqlalchemy.orm import joinedload
from collections import defaultdict
from libs.db_util import Item, Database  # 実際のモジュールパスに合わせてください。

class Test:
    def execute(self):
        self.session = Database.setup()
        self.test_count_item_per_group()
        self.session.close()

    def test_count_item_per_group(self):
        items = self.session.query(Item).options(joinedload(Item.images)).all()

        # siteとcategoryごとにアイテムをグループ化
        grouped_items = defaultdict(list)
        for item in items:
            grouped_items[(item.site, item.category)].append(item)

        # 各グループに対して処理を実行
        for (site, category), group_items in grouped_items.items():
            group_item_count = len(group_items)
            print(f"Group (site={site}, category={category}) のItemテーブルの件数: {group_item_count}")

            # Excelファイルを開く
            from datetime import datetime
            file_name = '【' + datetime.now().strftime("%Y年%m月") + f'分納品物】フリマサイト出品投稿一覧_{site}_{category}.xlsx'
            workbook = load_workbook(file_name)
            sheet = workbook['Infrigement Report']

            # C列の5行目から下に向かって値が入っている行数をカウント
            filled_rows_count = sum(1 for row in sheet.iter_rows(min_row=5, min_col=3, max_col=3, values_only=True) if row[0])
            print(f"Excelファイル '{file_name}' のC列に値が入っている行数: {filled_rows_count}")
