from openpyxl import load_workbook
from libs.db_util import HrInfo, Database

class Writer:
    def execute(self):
        self.session = Database.setup()
        items = self.session.query(HrInfo).all()
        self.write(items)

        self.session.close()

    def write(self, data):
        print(f'Excelファイル出力中')

        # テンプレートのExcelファイルを開く
        workbook = load_workbook('input/template.xlsx')
        sheet = workbook['日経人事情報リスト']

        row = 1  # ヘッダー行を除いた開始行
        for item in data:
            sheet.cell(row=row, column=3, value=item.title)
            sheet.cell(row=row, column=4, value=item.date)
            sheet.cell(row=row, column=5, value=item.body)
            row += 1

        # Excelファイルに保存
        workbook.save('日経人事情報リスト.xlsx')
